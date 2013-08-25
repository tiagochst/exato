from PyQt4 import QtGui, QtCore
from PyQt4.QtCore import *
from PyQt4.QtGui import *
import sys
import re
from openpyxl import load_workbook

def find(exatoPath,vestibularPath,function):
    # Alunos
    exato = []
    exato = readStudents(str(exatoPath),2013)
    # Vestibulares
    aprovados = []
    aprovados = function(str(vestibularPath))
    # Compare lists
    cmpNames(exato,aprovados)

# Remove accents from names and upper case it
def rmAccents(data):
    import unicodedata
    import string
    return ''.join(x for x in unicodedata.normalize('NFKD', data) if (x in string.ascii_letters) | (x == " ")).upper()


def showResult(nbOfnames, names):

    if(nbOfnames > 0):
        QMessageBox.information(None,"Lista aprovados",names)
    else:
        QMessageBox.information(None,"Lista aprovados","Nenhum aprovado")

def cmpNames(students, vestibular):
    msgAprovados = []
    for name in students:
        for aprovados in vestibular:
            if(name == aprovados):
                msgAprovados.append(name+"\n")
                break
    msg = ''.join(msgAprovados)
    showResult(len(msgAprovados),msg)

def readUnesp(fileName):
    # transform unesp pdf to txt
    from pdf2txt import pdf2txt
    import os
    arg = []
    arg.append("-o")
    arg.append("unesp.txt")
    arg.append(fileName)
    pdf2txt(arg)
    
    unesp = []
    f = open('unesp.txt', 'r')
    for line in f:
        if(line.isupper()):
            unesp.append(line.rpartition("\n")[0])
    f.close()
    os.remove('unesp.txt')     
    return unesp

def readUfscar(fileName):
    # transform ufscar pdf to txt
    from pdf2txt import pdf2txt
    import os

    arg = []
    arg.append("-o")
    arg.append("ufscar.txt")
    arg.append(fileName)
    pdf2txt(arg)
    
    ufscar = []
    f = open('ufscar.txt', 'r')
    for line in f:
        if(line.isupper()):
            ufscar.append(line.rpartition("\n")[0])
    f.close()
    os.remove('ufscar.txt')     
    return ufscar

def readUsp(fileName):
    usp = []
    f = open(fileName, 'r')
    for line in f:
        if(line.isupper()):
            if(line[0]!= "-"):
                line = str(re.split(r"([0-9]+)", line, flags=0)[0])
                usp.append(line.rstrip())
    f.close()
    return usp

def readUnicamp(fileName):
    unicamp = []
    f = open(fileName, 'r')
    for line in f:
        line = re.split(r"([0-9]+)", line, flags=0)
        if(len(line) > 2):
            line = line[2].rpartition("(")[0]
            unicamp.append(rmAccents(unicode(line.rstrip().lstrip())))
    f.close()
    
    return unicamp

# Reading a xlsx
def readStudents(fileName,year):
    wb = load_workbook(fileName)
    sheet = wb.get_sheet_by_name(name = str(year))
    
    students = []
    line = 1
    while True: 
        name = sheet.cell('A'+str(line)).value
        if(name):
            students.append(rmAccents(name)) 
            line = line + 1
        else: 
            break
    return students
